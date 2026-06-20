r"""
ProjectAnalyzer.py - CRP Project Intelligence Engine
======================================================

Scans PycharmProjects folders, SSMS SQL scripts, and Windows Task Scheduler
to build a comprehensive inventory of CRP's development ecosystem. Generates
per-project prompts for Claude to write narrative synopses, plus a cross-project
Excel workbook showing how everything connects.

Built for PLM onboarding — gives new team members (and AI assistants) a complete
picture of what exists, what each piece does, and how they fit together.

Lives in: CRPUtils folder
Output:   Results/ProjectAnalyzer_YYYY-MM-DD/

Reuses scanning logic from GPTProjectUploadGUI.py and GPTProjectUploadGUI_SQL.py.

Author: Pat Yearick
Created: April 2026
"""

import os
import re
import json
import logging
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Set
from dataclasses import dataclass, field
import sys

# =============================================================================
# CONFIGURATION
# =============================================================================

LOG_FILE = r"C:\Logs\ProjectAnalyzer.log"
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "ProjectAnalyzer_config.json")

# Default paths
DEFAULT_PYCHARM_ROOT = (
    r"C:\Users\pyearick.CRP\OneDrive - CRP Industries Inc"
    r"\CRPAF\PycharmProjects"
)
DEFAULT_SSMS_ROOT = (
    r"C:\Users\pyearick.CRP\OneDrive - CRP Industries Inc"
    r"\CRPAF\SQL Server Management Studio"
)
DEFAULT_OUTPUT_ROOT = (
    r"C:\Users\pyearick.CRP\OneDrive - CRP Industries Inc"
    r"\CRPAF\Results"
)

# Folders to skip when discovering projects and walking files
SKIP_FOLDERS = {
    '.idea', '.git', '.venv', 'venv', '__pycache__', 'node_modules',
    'CommitsGH', '.ipynb_checkpoints', '.venvBISQL001', 'Archive',
    'PunchlistReview',
}

# File extensions to scan
CODE_EXTENSIONS = {'.py', '.ps1', '.bat'}
DOC_EXTENSIONS = {'.md'}
SQL_EXTENSIONS = {'.sql'}
ALL_EXTENSIONS = CODE_EXTENSIONS | DOC_EXTENSIONS | SQL_EXTENSIONS

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ScriptInfo:
    """Metadata for a single script file."""
    filename: str
    relative_path: str
    full_path: str
    extension: str
    size: int
    created: str
    modified: str
    line_count: int
    imports: List[str] = field(default_factory=list)
    tables_read: List[str] = field(default_factory=list)
    tables_written: List[str] = field(default_factory=list)
    cross_project_imports: List[str] = field(default_factory=list)
    is_entry_point: bool = False
    has_argparse: bool = False
    docstring: str = ""


@dataclass
class DocInfo:
    """Metadata for a markdown or documentation file."""
    filename: str
    relative_path: str
    full_path: str
    size: int
    modified: str
    line_count: int
    doc_type: str  # 'decision_log', 'punchlist', 'readme', 'documentation'
    first_heading: str = ""
    preview: str = ""  # First ~500 chars


@dataclass
class SSMSScript:
    """Metadata for a SQL Server Management Studio script."""
    filename: str
    full_path: str
    size: int
    modified: str
    line_count: int
    tables_referenced: List[str] = field(default_factory=list)
    schema_objects: List[str] = field(default_factory=list)
    dependencies: List[str] = field(default_factory=list)


@dataclass
class ScheduledJob:
    """Metadata for a Windows Task Scheduler job."""
    task_name: str
    task_path: str
    state: str
    last_run: str
    next_run: str
    trigger_info: str
    action_command: str
    action_arguments: str
    author: str = ""
    description: str = ""


@dataclass
class NssmService:
    """Metadata for an NSSM-managed Windows service."""
    service_name: str
    display_name: str
    status: str           # Running, Stopped, Paused
    startup_type: str     # Auto, Manual, Disabled
    application: str      # Path to the executable
    app_directory: str    # Working directory
    app_arguments: str    # Command-line arguments
    related_project: str = ""  # Matched project name


@dataclass
class ProjectInventory:
    """Complete inventory for one project."""
    project_name: str
    project_path: str
    scripts: List[ScriptInfo] = field(default_factory=list)
    docs: List[DocInfo] = field(default_factory=list)
    sql_files: List[ScriptInfo] = field(default_factory=list)
    all_tables_read: Set[str] = field(default_factory=set)
    all_tables_written: Set[str] = field(default_factory=set)
    all_cross_imports: Set[str] = field(default_factory=set)
    entry_points: List[str] = field(default_factory=list)
    total_lines: int = 0
    total_files: int = 0


@dataclass
class CrossReference:
    """Cross-project analysis results."""
    table_readers: Dict[str, List[str]] = field(default_factory=dict)
    table_writers: Dict[str, List[str]] = field(default_factory=dict)
    import_graph: Dict[str, Set[str]] = field(default_factory=dict)
    shared_tables: Dict[str, Dict] = field(default_factory=dict)


# =============================================================================
# CONFIG PERSISTENCE
# =============================================================================

def load_config() -> dict:
    """Load persisted settings."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            pass
    return {
        'pycharm_root': DEFAULT_PYCHARM_ROOT,
        'ssms_root': DEFAULT_SSMS_ROOT,
        'output_root': DEFAULT_OUTPUT_ROOT,
        'include_ssms': True,
        'selected_projects': [],
        'selected_jobs': [],
    }


def save_config(config: dict):
    """Persist settings between runs."""
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=2, default=str)
    except Exception as e:
        logger.warning(f"Could not save config: {e}")


# =============================================================================
# PROJECT DISCOVERY
# =============================================================================

def discover_projects(pycharm_root: str) -> List[str]:
    """
    Find all project folders (immediate children of PycharmProjects).
    Returns sorted list of folder names, excluding SKIP_FOLDERS.
    """
    projects = []
    root_path = Path(pycharm_root)
    if not root_path.exists():
        logger.error(f"PycharmProjects root not found: {pycharm_root}")
        return projects

    for child in sorted(root_path.iterdir()):
        if (child.is_dir()
                and child.name not in SKIP_FOLDERS
                and not child.name.startswith('.')):
            projects.append(child.name)

    logger.info(f"Discovered {len(projects)} project folders")
    return projects


# =============================================================================
# EXTRACTION HELPERS — enhanced from GPTProjectUploadGUI
# =============================================================================

def extract_imports(content: str) -> List[str]:
    """Extract import statements from Python code."""
    imports = []
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("import ") or stripped.startswith("from "):
            imports.append(stripped)
    return imports


def extract_cross_project_imports(imports: List[str], project_name: str) -> List[str]:
    """
    Identify imports that reference sibling projects.
    e.g., 'from CRPUtils.logging_utils import setup_logging' -> 'CRPUtils'
    e.g., sys.path.append patterns pointing to sibling folders
    """
    cross = set()
    for imp in imports:
        # Direct imports: from ProjectName.module import ...
        match = re.match(r'from\s+(\w+)\.', imp)
        if match:
            ref = match.group(1)
            if ref != project_name and ref not in {'os', 'sys', 'datetime', 'pathlib',
                                                    'typing', 'collections', 'dataclasses',
                                                    'tkinter', 'json', 're', 'logging',
                                                    'subprocess', 'shutil', 'glob', 'io',
                                                    'hashlib', 'zipfile', 'xml', 'html',
                                                    'enum', 'abc', 'functools', 'itertools',
                                                    'csv', 'math', 'time', 'threading',
                                                    'concurrent', 'urllib', 'http',
                                                    'openpyxl', 'pandas', 'numpy', 'pyodbc',
                                                    'sqlalchemy', 'requests', 'PIL',
                                                    'matplotlib', 'seaborn', 'scipy',
                                                    'openai', 'dotenv', 'reportlab',
                                                    'pptx', 'selenium', 'undetected_chromedriver',
                                                    'bs4', 'pymongo', 'pyautogui',
                                                    'win32com', 'warnings', 'argparse',
                                                    'ast', 'textwrap', 'tempfile',
                                                    'traceback', 'socket', 'struct',
                                                    'base64', 'uuid', 'copy', 'string',
                                                    'difflib', 'inspect', 'signal',
                                                    'multiprocessing', 'queue'}:
                cross.add(ref)
    return sorted(cross)


# Identifiers that are databases or schemas, never actual tables. Used to keep
# qualified names like [BIWarehouse].[dbo].[Foo] from emitting the db/schema as
# a phantom "table" that every project appears to share.
_NON_TABLE_TOKENS = {
    'dbo', 'sys', 'information_schema', 'guest',
    'biwarehouse', 'crpaf', 'pricebooks', 'master', 'tempdb', 'model', 'msdb',
}
_SQL_NOISE = {
    'select', 'where', 'and', 'or', 'not', 'null', 'on', 'as',
    'inner', 'left', 'right', 'outer', 'cross', 'join', 'group', 'order',
}

# A table reference must carry a real SQL signal so we never mistake Python's
# own `from X import Y` for a table read. Two accepted shapes, both capturing
# the LAST (table) segment and discarding any db/schema prefix:
#   _BRACKETED : [db].[schema].[Table] / [schema].[Table] / [Table]  (table is bracketed)
#   _DBO_QUAL  : db.dbo.Table / dbo.Table                            (dbo-qualified)
# Bare unbracketed words after FROM/JOIN are NOT matched - that is what caused
# `from collections import ...` to be logged as a table.
_BRACKETED = r'(?:\[\w+\]\.){0,2}\[(\w+)\]'
_DBO_QUAL = r'(?:\w+\.)?dbo\.(\w+)'


def _keep_table(name: str) -> bool:
    low = name.lower()
    return low not in _SQL_NOISE and low not in _NON_TABLE_TOKENS


def _find_tables(content, keyword):
    """Collect table names after `keyword` (e.g. 'FROM|JOIN'), both SQL shapes."""
    found = set()
    for tail in (_BRACKETED, _DBO_QUAL):
        for match in re.finditer(rf'(?:{keyword})\s+' + tail, content, re.IGNORECASE):
            t = match.group(1)
            if _keep_table(t):
                found.add(t)
    return found


def extract_tables_read(content: str) -> List[str]:
    """
    Extract SQL table names that are READ from.
    Looks for: FROM/JOIN [db].[schema].[table] or dbo.table, plus read_sql_table.
    """
    tables = _find_tables(content, r'FROM|JOIN')

    # pandas: read_sql_table("TableName", ...)
    for match in re.finditer(r'read_sql_table\(\s*["\'](\w+)["\']', content):
        tables.add(match.group(1))

    return sorted(tables)


def extract_tables_written(content: str) -> List[str]:
    """
    Extract SQL table names that are WRITTEN to.
    Looks for: INSERT INTO, UPDATE, to_sql, MERGE INTO, TRUNCATE TABLE, DELETE FROM.
    """
    tables = _find_tables(
        content, r'INSERT\s+INTO|MERGE\s+INTO|DELETE\s+FROM|UPDATE|TRUNCATE\s+TABLE')
    for match in re.finditer(r'\.to_sql\(\s*["\'](\w+)["\']', content):
        tables.add(match.group(1))
    return sorted(tables)


def extract_docstring(content: str) -> str:
    """Extract the module-level docstring (first triple-quoted block)."""
    # Match r""" or """ at the start of file (possibly after comments/whitespace)
    match = re.match(r'\s*(?:r)?(?:"""(.*?)"""|\'\'\'(.*?)\'\'\')',
                     content, re.DOTALL)
    if match:
        doc = match.group(1) or match.group(2)
        # Clean up and truncate
        doc = doc.strip()
        if len(doc) > 1500:
            doc = doc[:1500] + "..."
        return doc
    return ""


def detect_entry_point(content: str) -> bool:
    """Check if script has if __name__ == '__main__' block."""
    return bool(re.search(r'if\s+__name__\s*==\s*["\']__main__["\']', content))


def detect_argparse(content: str) -> bool:
    """Check if script uses argparse for CLI arguments."""
    return 'argparse' in content or 'ArgumentParser' in content


def classify_markdown(filename: str, content: str) -> str:
    """Classify a markdown file by type."""
    name_lower = filename.lower()

    if 'decisionlog' in name_lower or 'decision_log' in name_lower:
        return 'decision_log'
    if 'punchlist' in name_lower:
        return 'punchlist'
    if name_lower in ('readme.md', 'readme'):
        return 'readme'

    # Check content for decision log indicators
    content_lower = content[:2000].lower()
    if 'decision' in content_lower and ('log' in content_lower or 'rationale' in content_lower):
        return 'decision_log'

    return 'documentation'


def extract_first_heading(content: str) -> str:
    """Extract the first markdown heading."""
    match = re.search(r'^#\s+(.+)', content, re.MULTILINE)
    return match.group(1).strip() if match else ""


# =============================================================================
# SQL EXTRACTION — from GPTProjectUploadGUI_SQL
# =============================================================================

def extract_sql_dependencies(content: str) -> List[str]:
    """Extract dependencies from a T-SQL file."""
    deps = []
    for line in content.splitlines():
        stripped = line.strip().upper()
        if stripped.startswith("USE ") or stripped.startswith("EXEC ") or stripped.startswith("EXECUTE "):
            deps.append(line.strip())
        elif "SERVER" in stripped and any(kw in stripped for kw in ["OPENQUERY", "OPENDATASOURCE"]):
            deps.append(line.strip())
    return deps


def extract_schema_objects(content: str) -> List[str]:
    """Extract schema objects (CREATE TABLE, VIEW, PROC, etc.) from T-SQL."""
    objects = []
    keywords = [
        "CREATE TABLE", "CREATE VIEW", "CREATE PROCEDURE", "CREATE FUNCTION",
        "CREATE TRIGGER", "CREATE INDEX", "CREATE SCHEMA", "CREATE TYPE",
        "ALTER TABLE", "ALTER VIEW", "ALTER PROCEDURE", "ALTER FUNCTION"
    ]
    for line in content.splitlines():
        stripped = line.strip().upper()
        if any(kw in stripped for kw in keywords):
            objects.append(line.strip())
    return objects


# =============================================================================
# PROJECT SCANNING
# =============================================================================

def scan_project(project_path: str) -> ProjectInventory:
    """
    Walk a single project folder and build a complete inventory.
    Scans .py, .ps1, .bat, .md, and .sql files.
    """
    project_name = os.path.basename(project_path)
    inventory = ProjectInventory(project_name=project_name, project_path=project_path)

    logger.info(f"Scanning project: {project_name}")

    for root, dirs, files in os.walk(project_path):
        # Prune excluded directories
        dirs[:] = [d for d in dirs if d.lower() not in {s.lower() for s in SKIP_FOLDERS}]

        for filename in sorted(files):
            filepath = os.path.join(root, filename)
            ext = os.path.splitext(filename)[1].lower()

            if ext not in ALL_EXTENSIONS:
                continue

            try:
                with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
            except Exception as e:
                logger.warning(f"  Could not read {filepath}: {e}")
                continue

            relative = os.path.relpath(filepath, project_path)
            stat = os.stat(filepath)
            line_count = len(content.splitlines())

            # --- Code files (.py, .ps1, .bat) ---
            if ext in CODE_EXTENSIONS:
                imports = extract_imports(content) if ext == '.py' else []
                cross_imports = extract_cross_project_imports(imports, project_name) if ext == '.py' else []
                tables_r = extract_tables_read(content)
                tables_w = extract_tables_written(content)
                is_entry = detect_entry_point(content) if ext == '.py' else False
                has_args = detect_argparse(content) if ext == '.py' else False
                docstring = extract_docstring(content) if ext == '.py' else ""

                script = ScriptInfo(
                    filename=filename,
                    relative_path=relative,
                    full_path=filepath,
                    extension=ext,
                    size=stat.st_size,
                    created=datetime.fromtimestamp(stat.st_ctime).isoformat(),
                    modified=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    line_count=line_count,
                    imports=imports,
                    tables_read=tables_r,
                    tables_written=tables_w,
                    cross_project_imports=cross_imports,
                    is_entry_point=is_entry,
                    has_argparse=has_args,
                    docstring=docstring,
                )
                inventory.scripts.append(script)
                inventory.all_tables_read.update(tables_r)
                inventory.all_tables_written.update(tables_w)
                inventory.all_cross_imports.update(cross_imports)
                if is_entry:
                    inventory.entry_points.append(filename)

            # --- SQL files ---
            elif ext in SQL_EXTENSIONS:
                tables_r = extract_tables_read(content)
                tables_w = extract_tables_written(content)

                sql_info = ScriptInfo(
                    filename=filename,
                    relative_path=relative,
                    full_path=filepath,
                    extension=ext,
                    size=stat.st_size,
                    created=datetime.fromtimestamp(stat.st_ctime).isoformat(),
                    modified=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    line_count=line_count,
                    tables_read=tables_r,
                    tables_written=tables_w,
                    docstring="",
                )
                inventory.sql_files.append(sql_info)
                inventory.all_tables_read.update(tables_r)
                inventory.all_tables_written.update(tables_w)

            # --- Markdown files ---
            elif ext in DOC_EXTENSIONS:
                doc_type = classify_markdown(filename, content)
                first_heading = extract_first_heading(content)
                preview = content[:500].strip()
                if len(content) > 500:
                    preview += "..."

                doc = DocInfo(
                    filename=filename,
                    relative_path=relative,
                    full_path=filepath,
                    size=stat.st_size,
                    modified=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    line_count=line_count,
                    doc_type=doc_type,
                    first_heading=first_heading,
                    preview=preview,
                )
                inventory.docs.append(doc)

            inventory.total_lines += line_count
            inventory.total_files += 1

    logger.info(f"  {project_name}: {len(inventory.scripts)} scripts, "
                f"{len(inventory.sql_files)} SQL, {len(inventory.docs)} docs, "
                f"{inventory.total_lines} total lines")

    return inventory


# =============================================================================
# SSMS SCANNING
# =============================================================================

def scan_ssms_folder(ssms_root: str) -> List[SSMSScript]:
    """
    Scan the SSMS folder for .sql scripts. Flat directory structure.
    These are historical/foundation scripts that built tables and procs.
    """
    scripts = []
    ssms_path = Path(ssms_root)

    if not ssms_path.exists():
        logger.warning(f"SSMS folder not found: {ssms_root}")
        return scripts

    for filepath in sorted(ssms_path.glob("*.sql")):
        try:
            content = filepath.read_text(encoding='utf-8', errors='replace')
            stat = filepath.stat()

            ssms = SSMSScript(
                filename=filepath.name,
                full_path=str(filepath),
                size=stat.st_size,
                modified=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                line_count=len(content.splitlines()),
                tables_referenced=extract_tables_read(content) + extract_tables_written(content),
                schema_objects=extract_schema_objects(content),
                dependencies=extract_sql_dependencies(content),
            )
            scripts.append(ssms)
        except Exception as e:
            logger.warning(f"  Could not read SSMS script {filepath.name}: {e}")

    # Also check Automation subfolder
    auto_path = ssms_path / "Automation"
    if auto_path.exists():
        for filepath in sorted(auto_path.glob("*.sql")):
            try:
                content = filepath.read_text(encoding='utf-8', errors='replace')
                stat = filepath.stat()
                ssms = SSMSScript(
                    filename=f"Automation/{filepath.name}",
                    full_path=str(filepath),
                    size=stat.st_size,
                    modified=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    line_count=len(content.splitlines()),
                    tables_referenced=extract_tables_read(content) + extract_tables_written(content),
                    schema_objects=extract_schema_objects(content),
                    dependencies=extract_sql_dependencies(content),
                )
                scripts.append(ssms)
            except Exception as e:
                logger.warning(f"  Could not read SSMS script {filepath.name}: {e}")

    logger.info(f"SSMS scan: {len(scripts)} SQL scripts found")
    return scripts


# =============================================================================
# TASK SCHEDULER SCANNING
# =============================================================================

def scan_scheduled_tasks(author_filter: str = "CRP") -> List[ScheduledJob]:
    """
    Query Windows Task Scheduler via PowerShell.
    Pre-filters to tasks authored by CRP accounts (CRP\\PYEARICK, CRP\\BI-SQL001$, etc.)
    to exclude system noise (Windows Defender, OneDrive, etc.).

    Args:
        author_filter: Only include tasks where Author starts with this prefix.
                       Default 'CRP' matches all CRP domain accounts.
                       Pass '' to get all tasks (unfiltered).
    """
    jobs = []

    ps_script = r"""
    $authorPrefix = '""" + author_filter + r"""'
    $tasks = Get-ScheduledTask
    if ($authorPrefix -ne '') {
        $tasks = $tasks | Where-Object {
            $_.Author -and $_.Author -like ($authorPrefix + '*')
        }
    } else {
        $tasks = $tasks | Where-Object { $_.State -ne 'Disabled' }
    }
    $results = @()
    foreach ($task in $tasks) {
        $action = $task.Actions | Select-Object -First 1
        $trigger = $task.Triggers | Select-Object -First 1

        $triggerType = 'N/A'
        if ($trigger) {
            try { $triggerType = $trigger.CimClass.CimClassName -replace 'MSFT_Task','' -replace 'Trigger','' }
            catch { $triggerType = 'Unknown' }
        }

        $cmd = 'N/A'
        if ($action -and $action.Execute) { $cmd = $action.Execute }

        $args = ''
        if ($action -and $action.Arguments) { $args = $action.Arguments }

        $desc = ''
        if ($task.Description) { $desc = $task.Description }

        $author = ''
        if ($task.Author) { $author = $task.Author }

        # Get run times separately with error handling
        $lastRun = 'N/A'
        $nextRun = 'N/A'
        try {
            $info = $task | Get-ScheduledTaskInfo -ErrorAction SilentlyContinue
            if ($info -and $info.LastRunTime -and $info.LastRunTime.Year -gt 1999) {
                $lastRun = $info.LastRunTime.ToString('yyyy-MM-dd HH:mm')
            }
            if ($info -and $info.NextRunTime -and $info.NextRunTime.Year -gt 1999) {
                $nextRun = $info.NextRunTime.ToString('yyyy-MM-dd HH:mm')
            }
        } catch {}

        $results += [PSCustomObject]@{
            TaskName    = $task.TaskName
            TaskPath    = $task.TaskPath
            State       = $task.State.ToString()
            Description = $desc
            Author      = $author
            LastRun     = $lastRun
            NextRun     = $nextRun
            TriggerType = $triggerType
            Command     = $cmd
            Arguments   = $args
        }
    }
    $results | ConvertTo-Json -Depth 3 -Compress
    """

    try:
        logger.info("Querying Windows Task Scheduler...")
        result = subprocess.run(
            ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass', '-Command', ps_script],
            capture_output=True, text=True, timeout=120,
            encoding='utf-8', errors='replace'
        )

        if result.returncode != 0:
            logger.warning(f"Task Scheduler query error (rc={result.returncode}): "
                           f"{result.stderr[:500]}")
            # Still try to parse stdout — PS sometimes writes warnings to stderr
            # but still produces valid JSON output

        stdout = result.stdout.strip()
        if not stdout:
            logger.warning("Task Scheduler returned empty output")
            if result.stderr:
                logger.warning(f"stderr: {result.stderr[:500]}")
            return jobs

        data = json.loads(stdout)
        if isinstance(data, dict):
            data = [data]  # Single task comes back as dict, not list

        for item in data:
            job = ScheduledJob(
                task_name=item.get('TaskName', ''),
                task_path=item.get('TaskPath', ''),
                state=item.get('State', ''),
                last_run=item.get('LastRun', 'N/A'),
                next_run=item.get('NextRun', 'N/A'),
                trigger_info=item.get('TriggerType', 'N/A'),
                action_command=item.get('Command', ''),
                action_arguments=item.get('Arguments', ''),
                author=item.get('Author', ''),
                description=item.get('Description', ''),
            )
            jobs.append(job)

        logger.info(f"Task Scheduler: {len(jobs)} active tasks found")

    except subprocess.TimeoutExpired:
        logger.error("Task Scheduler query timed out (120s) — "
                     "try running Get-ScheduledTask in PowerShell to verify access")
    except json.JSONDecodeError as e:
        logger.error(f"Could not parse Task Scheduler JSON: {e}")
        # Log first 500 chars of output for debugging
        if result.stdout:
            logger.error(f"Raw output (first 500): {result.stdout[:500]}")
    except Exception as e:
        logger.error(f"Task Scheduler scan failed: {e}")

    return jobs


# =============================================================================
# NSSM SERVICE SCANNING
# =============================================================================

def _nssm_get(nssm_path: str, service_name: str, parameter: str) -> str:
    """Query a single NSSM parameter. Handles UTF-16-LE output."""
    try:
        result = subprocess.run(
            [nssm_path, 'get', service_name, parameter],
            capture_output=True, timeout=10
        )
        # NSSM outputs UTF-16-LE — decode as bytes, strip nulls
        raw = result.stdout
        try:
            value = raw.decode('utf-16-le').strip().strip('\x00')
        except (UnicodeDecodeError, ValueError):
            value = raw.decode('utf-8', errors='replace').strip().strip('\x00')
        return value
    except Exception:
        return ''


def scan_nssm_services(pycharm_root: str) -> List[NssmService]:
    """
    Scan for NSSM-managed services. Queries each discovered NSSM service
    for its application path, arguments, and working directory.

    The NSSM executable lives in the PycharmProjects directory.
    """
    services = []

    # Find nssm.exe — lives in nlp-sql-in-a-box project directory
    nssm_path = None
    search_paths = [
        os.path.join(pycharm_root, 'nlp-sql-in-a-box', 'nssm.exe'),
        os.path.join(pycharm_root, 'nlp-sql-in-a-box', 'nssm', 'win64', 'nssm.exe'),
        os.path.join(pycharm_root, 'nssm.exe'),
        os.path.join(pycharm_root, 'nssm', 'nssm.exe'),
        os.path.join(pycharm_root, 'nssm', 'win64', 'nssm.exe'),
    ]
    for candidate in search_paths:
        if os.path.exists(candidate):
            nssm_path = candidate
            break

    if not nssm_path:
        logger.info("NSSM executable not found — skipping service scan")
        return services

    logger.info(f"Scanning NSSM services using: {nssm_path}")

    # Get list of all NSSM-managed services
    # NSSM doesn't have a native "list all" command, so we query Windows
    # services and check which ones are managed by NSSM
    try:
        ps_script = r"""
        Get-WmiObject Win32_Service |
            Where-Object { $_.PathName -like '*nssm*' } |
            Select-Object Name, DisplayName, State, StartMode |
            ConvertTo-Json -Compress
        """
        result = subprocess.run(
            ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass', '-Command', ps_script],
            capture_output=True, text=True, timeout=30,
            encoding='utf-8', errors='replace'
        )

        stdout = result.stdout.strip()
        if not stdout:
            logger.info("No NSSM services found")
            return services

        data = json.loads(stdout)
        if isinstance(data, dict):
            data = [data]

        for item in data:
            svc_name = item.get('Name', '')
            if not svc_name:
                continue

            # Query NSSM for details
            application = _nssm_get(nssm_path, svc_name, 'Application')
            app_dir = _nssm_get(nssm_path, svc_name, 'AppDirectory')
            app_args = _nssm_get(nssm_path, svc_name, 'AppParameters')

            # Try to match to a project folder
            related = ''
            if app_dir:
                # Check if AppDirectory is inside a project folder
                try:
                    rel = os.path.relpath(app_dir, pycharm_root)
                    if not rel.startswith('..'):
                        related = rel.split(os.sep)[0]
                except ValueError:
                    pass

            svc = NssmService(
                service_name=svc_name,
                display_name=item.get('DisplayName', svc_name),
                status=item.get('State', 'Unknown'),
                startup_type=item.get('StartMode', 'Unknown'),
                application=application,
                app_directory=app_dir,
                app_arguments=app_args,
                related_project=related,
            )
            services.append(svc)

        logger.info(f"NSSM scan: {len(services)} services found")

    except Exception as e:
        logger.error(f"NSSM service scan failed: {e}")

    return services


# =============================================================================
# CROSS-PROJECT ANALYSIS
# =============================================================================

def build_cross_reference(inventories: Dict[str, ProjectInventory]) -> CrossReference:
    """
    Analyze all projects together to find shared tables,
    import dependencies, and connection points.
    """
    xref = CrossReference()

    for proj_name, inv in inventories.items():
        # Track table readers
        for table in inv.all_tables_read:
            xref.table_readers.setdefault(table, []).append(proj_name)

        # Track table writers
        for table in inv.all_tables_written:
            xref.table_writers.setdefault(table, []).append(proj_name)

        # Track import graph - only edges to projects actually in the roster.
        # (A denylist upstream can't catch every pip package, so filter here:
        #  an import counts as cross-project only if it resolves to a real one.)
        if inv.all_cross_imports:
            known = sorted(d for d in inv.all_cross_imports if d in inventories)
            if known:
                xref.import_graph[proj_name] = known

    # Build shared tables map (tables touched by 2+ projects)
    all_tables = set(xref.table_readers.keys()) | set(xref.table_writers.keys())
    for table in all_tables:
        readers = set(xref.table_readers.get(table, []))
        writers = set(xref.table_writers.get(table, []))
        all_projects = readers | writers
        if len(all_projects) > 1:
            xref.shared_tables[table] = {
                'readers': sorted(readers),
                'writers': sorted(writers),
                'all_projects': sorted(all_projects),
            }

    logger.info(f"Cross-reference: {len(xref.shared_tables)} shared tables, "
                f"{len(xref.import_graph)} projects with cross-imports")

    return xref


# =============================================================================
# SSMS-TO-PROJECT MATCHING
# =============================================================================

def match_ssms_to_projects(ssms_scripts: List[SSMSScript],
                           inventories: Dict[str, ProjectInventory]) -> Dict[str, List[str]]:
    """
    Match SSMS SQL scripts to projects by overlapping table references.
    Returns: { ssms_filename: [project_names] }
    """
    matches = {}
    for ssms in ssms_scripts:
        ssms_tables = set(ssms.tables_referenced)
        if not ssms_tables:
            continue

        related = []
        for proj_name, inv in inventories.items():
            proj_tables = inv.all_tables_read | inv.all_tables_written
            overlap = ssms_tables & proj_tables
            if overlap:
                related.append(proj_name)

        if related:
            matches[ssms.filename] = sorted(related)

    return matches


# =============================================================================
# PROMPT GENERATION
# =============================================================================

def generate_project_prompt(inventory: ProjectInventory,
                            xref: CrossReference,
                            selected_jobs: List[ScheduledJob],
                            nssm_services: List[NssmService],
                            ssms_scripts: List[SSMSScript],
                            ssms_matches: Dict[str, List[str]]) -> str:
    """
    Generate a prompt for Claude in a specific project chat.
    The prompt contains the full inventory + context so Claude can
    write a narrative synopsis.
    """
    proj = inventory.project_name
    lines = []

    lines.append(f"# Project Analysis Request: {proj}")
    lines.append("")
    lines.append("An automated scan of this project has been completed. Below is the full")
    lines.append("inventory plus cross-project context. Please write a comprehensive synopsis")
    lines.append(f"and save it as `{proj}_Synopsis.md`.")
    lines.append("")

    # --- Section 1: Script Inventory ---
    lines.append("## SCRIPT INVENTORY")
    lines.append(f"Total: {len(inventory.scripts)} code files, "
                 f"{len(inventory.sql_files)} SQL files, "
                 f"{len(inventory.docs)} documentation files, "
                 f"{inventory.total_lines} total lines")
    lines.append("")

    for script in inventory.scripts:
        lines.append(f"### {script.filename}")
        lines.append(f"- Lines: {script.line_count} | Modified: {script.modified[:10]}")
        if script.is_entry_point:
            lines.append(f"- **Entry point** (runnable script)")
        if script.has_argparse:
            lines.append(f"- Has CLI arguments (argparse)")
        if script.docstring:
            lines.append(f"- Purpose: {script.docstring}")
        if script.tables_read:
            lines.append(f"- Reads from: {', '.join(script.tables_read)}")
        if script.tables_written:
            lines.append(f"- Writes to: {', '.join(script.tables_written)}")
        if script.cross_project_imports:
            lines.append(f"- Imports from: {', '.join(script.cross_project_imports)}")
        lines.append("")

    # --- Section 2: SQL files in project ---
    if inventory.sql_files:
        lines.append("## SQL FILES IN PROJECT")
        for sql in inventory.sql_files:
            lines.append(f"- {sql.filename} ({sql.line_count} lines, modified {sql.modified[:10]})")
            if sql.tables_read:
                lines.append(f"  Reads: {', '.join(sql.tables_read)}")
            if sql.tables_written:
                lines.append(f"  Writes: {', '.join(sql.tables_written)}")
        lines.append("")

    # --- Section 3: Documentation ---
    if inventory.docs:
        lines.append("## DOCUMENTATION FILES")
        for doc in inventory.docs:
            label = doc.doc_type.replace('_', ' ').title()
            lines.append(f"### {doc.filename} [{label}]")
            if doc.first_heading:
                lines.append(f"Heading: {doc.first_heading}")
            lines.append(f"Preview:")
            lines.append(f"```")
            lines.append(doc.preview)
            lines.append(f"```")
            lines.append("")

    # --- Section 4: Cross-Project Connections ---
    lines.append("## CROSS-PROJECT CONNECTIONS")
    _imports_clean = xref.import_graph.get(proj, [])
    if _imports_clean:
        lines.append(f"This project imports from: {', '.join(sorted(_imports_clean))}")

    # Who imports from us?
    importers = [p for p, deps in xref.import_graph.items()
                 if proj in deps and p != proj]
    if importers:
        lines.append(f"These projects import from {proj}: {', '.join(sorted(importers))}")

    # Shared tables
    proj_tables = inventory.all_tables_read | inventory.all_tables_written
    relevant_shared = {t: info for t, info in xref.shared_tables.items()
                       if t in proj_tables}
    if relevant_shared:
        lines.append("")
        lines.append("### Shared Tables (used by multiple projects)")
        for table, info in sorted(relevant_shared.items()):
            other_projects = [p for p in info['all_projects'] if p != proj]
            rw = []
            if proj in info.get('readers', []):
                rw.append('reads')
            if proj in info.get('writers', []):
                rw.append('writes')
            lines.append(f"- **{table}**: {proj} {'/'.join(rw)} | "
                         f"also used by: {', '.join(other_projects)}")
    lines.append("")

    # --- Section 5: Scheduled Jobs ---
    related_jobs = [j for j in selected_jobs
                    if proj.lower() in j.action_command.lower()
                    or proj.lower() in j.action_arguments.lower()]
    if related_jobs:
        lines.append("## SCHEDULED AUTOMATION")
        for job in related_jobs:
            lines.append(f"- **{job.task_name}** ({job.trigger_info})")
            lines.append(f"  Command: {job.action_command} {job.action_arguments}")
            lines.append(f"  Last run: {job.last_run} | Next: {job.next_run}")
        lines.append("")

    # --- Section 5b: NSSM Services ---
    related_nssm = [s for s in nssm_services
                    if s.related_project.lower() == proj.lower()
                    or proj.lower() in s.application.lower()
                    or proj.lower() in s.app_directory.lower()]
    if related_nssm:
        lines.append("## NSSM SERVICES (always-on)")
        lines.append("These Windows services run continuously via NSSM (not Task Scheduler):")
        for svc in related_nssm:
            lines.append(f"- **{svc.service_name}** — Status: {svc.status}, Startup: {svc.startup_type}")
            lines.append(f"  Application: {svc.application}")
            if svc.app_arguments:
                lines.append(f"  Arguments: {svc.app_arguments}")
            lines.append(f"  Working dir: {svc.app_directory}")
        lines.append("")
        lines.append("NOTE: NSSM services run their own internal schedules (cron-like loops,")
        lines.append("timed cycles, etc.). Please document the internal scheduling logic")
        lines.append("found in the scripts, not just the fact that the service runs.")
        lines.append("")

    # --- Section 6: Related SSMS Scripts ---
    related_ssms = [s for s in ssms_scripts
                    if s.filename in ssms_matches
                    and proj in ssms_matches[s.filename]]
    if related_ssms:
        lines.append("## RELATED SSMS SQL SCRIPTS (historical)")
        lines.append("These SQL scripts from the SSMS folder reference tables also used by this project:")
        for s in related_ssms[:20]:  # Cap at 20 to keep prompt manageable
            lines.append(f"- {s.filename} ({s.line_count} lines)")
            if s.schema_objects:
                lines.append(f"  Creates: {'; '.join(s.schema_objects[:5])}")
        lines.append("")

    # --- Section 7: Output instructions ---
    lines.append("## OUTPUT INSTRUCTIONS")
    lines.append(f"Please write `{proj}_Synopsis.md` following the EXACT format below.")
    lines.append("Consistency across all project synopses is critical — these will be read")
    lines.append("side-by-side by new team members. Follow the template precisely.")
    lines.append("")
    lines.append("### HEADER FORMAT")
    lines.append(f"```")
    lines.append(f"# {proj} — Project Synopsis")
    lines.append(f"_CRP Industries · CRPAF Team · [Month Year]_")
    lines.append(f"```")
    lines.append("")
    lines.append("### REQUIRED SECTIONS (use these exact headings)")
    lines.append("")
    lines.append("**## 1. Project Purpose**")
    lines.append("2-3 paragraphs. What problem does this solve? Why does it exist?")
    lines.append("Write for someone who has never seen this project.")
    lines.append("")
    lines.append("**## 2. History & Evolution**")
    lines.append("Structure around NAMED ARCHITECTURAL INFLECTION POINTS, not dates.")
    lines.append("Example: 'The Modular Pipeline Redesign', 'The Staging Table Pattern'.")
    lines.append("Each inflection point gets a bold name, a paragraph explaining what")
    lines.append("changed and WHY. Dates are secondary context, not primary organization.")
    lines.append("When a decision log exists, reference specific entry numbers inline")
    lines.append("(e.g., 'documented as Decision 012 in the Decision Log').")
    lines.append("If the project is mid-migration between platforms or architectures")
    lines.append("(e.g., Power Apps to Streamlit, monolith to modular), describe the")
    lines.append("current state and what remains on the legacy platform.")
    lines.append("")
    lines.append("**## 3. Script Inventory**")
    lines.append("Group scripts into NAMED FUNCTIONAL AREAS (not a flat list).")
    lines.append("Example areas: 'Core Analysis Engine', 'Web Scrapers', 'Reporting'.")
    lines.append("For each area, write a brief narrative intro, then for each script:")
    lines.append("  **`ScriptName.py`** (N lines) — Narrative paragraph explaining what")
    lines.append("  it does, what data it reads/writes, and its role in the pipeline.")
    lines.append("  Include key table names and cross-project imports inline.")
    lines.append("LINE COUNTS: Include (N lines) for projects with 15+ scripts where")
    lines.append("relative size helps understand architecture. For smaller projects,")
    lines.append("line counts are optional — omit if they add noise without insight.")
    lines.append("DEFAULT to narrative paragraphs for individual scripts.")
    lines.append("EXCEPTION: For functional areas with 5+ tightly related scripts that")
    lines.append("would be repetitive in paragraph form (e.g., a suite of scrapers or")
    lines.append("a numbered pipeline), a summary table within that area is acceptable.")
    lines.append("The area-level description must still be narrative.")
    lines.append("")
    lines.append("**## 4. Data Flow**")
    lines.append("Separate into 'Sources (Read)' and 'Destinations (Write)' sub-sections.")
    lines.append("If most tables share a single database (e.g., CRPAF.dbo), state that")
    lines.append("once at the top and use a two-column format: Table | What It Provides.")
    lines.append("Only use three columns (Database | Table | Purpose) when tables span")
    lines.append("multiple databases.")
    lines.append("Include any external services, APIs, or file-based data flows.")
    lines.append("For projects with 15+ tables: organize by pipeline phase or functional")
    lines.append("area rather than one flat list. Curate to the key tables and note")
    lines.append("'N additional tables used for intermediate processing' if needed.")
    lines.append("End with an 'End-to-End Pipeline' numbered list showing the full flow.")
    lines.append("")
    lines.append("**## 5. Cross-Project Dependencies**")
    lines.append("For each shared table or cross-project import, explain the RELATIONSHIP —")
    lines.append("not just that it's shared, but what data flows between projects and why.")
    lines.append("")
    lines.append("**## 6. Automation**")
    lines.append("Cover ALL automation layers that apply:")
    lines.append("  - **Task Scheduler jobs** — what runs, schedule, last/next run")
    lines.append("  - **NSSM services** — always-on services, what they host")
    lines.append("  - **Persistent web services** — Streamlit apps, Flask servers, API")
    lines.append("    endpoints. Include URL/port and how users access them.")
    lines.append("  - **Internal scheduling** — cron-like loops, timed cycles, schedule")
    lines.append("    configs within scripts (e.g., PMA_Orchestrator's agent schedule)")
    lines.append("For each: what runs, when, what triggers it, what it produces.")
    lines.append("Omit any category that doesn't apply to this project.")
    lines.append("")
    lines.append("**## 7. Known Issues & Technical Debt**")
    lines.append("Organize into three tiers:")
    lines.append("  - **Active Blockers** — things preventing features or causing failures")
    lines.append("  - **Open Issues** — known problems that need fixing but aren't blocking")
    lines.append("  - **Technical Debt** — code quality, cleanup, refactoring opportunities")
    lines.append("Be honest and specific. Name the scripts, the problems, the impact.")
    lines.append("Flag: duplicated code, hardcoded values, archived files not cleaned up,")
    lines.append("missing tests, unreliable automation, placeholder values.")
    lines.append("If an Archive folder exists in the project, note its presence and")
    lines.append("approximate contents but do not detail individual archived scripts.")
    lines.append("")
    lines.append("**## 8. PLM Relevance**")
    lines.append("Use bold subsection headers for each PLM dimension this project touches.")
    lines.append("Choose from (or add project-specific dimensions as needed):")
    lines.append("Product Discovery, Pricing Intelligence, Supplier Management,")
    lines.append("Catalog Quality, Inventory Optimization, Customer Intelligence,")
    lines.append("Competitive Intelligence, Operational Monitoring, Product Onboarding,")
    lines.append("Workflow Automation.")
    lines.append("Keep each subsection to 2-3 sentences. Be specific about HOW.")
    lines.append("")
    lines.append("### FOOTER FORMAT")
    lines.append(f"```")
    lines.append(f"---")
    lines.append(f"_Generated from automated project scan. [N] code files, [N] total lines.")
    lines.append(f"Last code change: [date] (`filename`).")
    lines.append(f"Last scheduled activity: [date] (omit if no scheduled automation).")
    lines.append(f"Decision log: [Yes/No]._")
    lines.append(f"```")
    lines.append("")
    lines.append("### STYLE RULES")
    lines.append("- NO summary statistics tables. The narrative IS the summary.")
    lines.append("- Use `backtick` formatting for script names, table names, and SQL objects.")
    lines.append("- Bold (**text**) for emphasis on key decisions and important findings.")
    lines.append("- Keep paragraphs to 3-5 sentences. Dense but readable.")
    lines.append("- If a decision log exists, include a '### Key Decisions' appendix")
    lines.append("  with a table: Date | Decision | Rationale | Impact.")
    lines.append("- If a punchlist or open issues document exists, reference it by name")
    lines.append("  and summarize the top 3-5 items. Do not reproduce the full list.")
    lines.append("")
    lines.append("Be specific. Use actual script names, table names, and data flows.")
    lines.append("Let's discuss approach before writing if anything is unclear.")

    return '\n'.join(lines)


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_workbook(inventories: Dict[str, ProjectInventory],
                    xref: CrossReference,
                    ssms_scripts: List[SSMSScript],
                    ssms_matches: Dict[str, List[str]],
                    selected_jobs: List[ScheduledJob],
                    nssm_services: List[NssmService],
                    output_path: str):
    """
    Export the cross-reference data to an Excel workbook.
    Tabs: Project Index, Table Map, Import Graph, Scheduled Jobs, NSSM Services, SSMS Scripts
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl not installed — skipping Excel export")
        return

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')

    def write_headers(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

    # --- Tab 1: Project Index ---
    ws1 = wb.active
    ws1.title = "Project Index"
    write_headers(ws1, ['Project', 'Scripts', 'SQL Files', 'Docs', 'Total Lines',
                        'Entry Points', 'Tables Read', 'Tables Written',
                        'Cross-Project Imports'])
    for row, (name, inv) in enumerate(sorted(inventories.items()), 2):
        ws1.cell(row=row, column=1, value=name)
        ws1.cell(row=row, column=2, value=len(inv.scripts))
        ws1.cell(row=row, column=3, value=len(inv.sql_files))
        ws1.cell(row=row, column=4, value=len(inv.docs))
        ws1.cell(row=row, column=5, value=inv.total_lines)
        ws1.cell(row=row, column=6, value=', '.join(inv.entry_points))
        ws1.cell(row=row, column=7, value=', '.join(sorted(inv.all_tables_read)))
        ws1.cell(row=row, column=8, value=', '.join(sorted(inv.all_tables_written)))
        ws1.cell(row=row, column=9, value=', '.join(sorted(inv.all_cross_imports)))

    # --- Tab 2: Shared Table Map ---
    ws2 = wb.create_sheet("Shared Tables")
    write_headers(ws2, ['Table', 'Written By', 'Read By', 'Total Projects'])
    for row, (table, info) in enumerate(sorted(xref.shared_tables.items()), 2):
        ws2.cell(row=row, column=1, value=table)
        ws2.cell(row=row, column=2, value=', '.join(info['writers']))
        ws2.cell(row=row, column=3, value=', '.join(info['readers']))
        ws2.cell(row=row, column=4, value=len(info['all_projects']))

    # --- Tab 3: All Tables (comprehensive) ---
    ws3 = wb.create_sheet("All Tables")
    write_headers(ws3, ['Table', 'Project', 'Direction', 'Script'])
    row = 2
    for proj_name, inv in sorted(inventories.items()):
        for script in inv.scripts + inv.sql_files:
            for t in script.tables_read:
                ws3.cell(row=row, column=1, value=t)
                ws3.cell(row=row, column=2, value=proj_name)
                ws3.cell(row=row, column=3, value='READ')
                ws3.cell(row=row, column=4, value=script.filename)
                row += 1
            for t in script.tables_written:
                ws3.cell(row=row, column=1, value=t)
                ws3.cell(row=row, column=2, value=proj_name)
                ws3.cell(row=row, column=3, value='WRITE')
                ws3.cell(row=row, column=4, value=script.filename)
                row += 1

    # --- Tab 4: Import Dependencies ---
    ws4 = wb.create_sheet("Import Graph")
    write_headers(ws4, ['Project', 'Imports From', 'Imported By'])
    all_proj_names = sorted(inventories.keys())
    for row, proj in enumerate(all_proj_names, 2):
        imports_from = ', '.join(sorted(xref.import_graph.get(proj, set())))
        imported_by = ', '.join(sorted(
            p for p, deps in xref.import_graph.items() if proj in deps and p != proj
        ))
        ws4.cell(row=row, column=1, value=proj)
        ws4.cell(row=row, column=2, value=imports_from)
        ws4.cell(row=row, column=3, value=imported_by)

    # --- Tab 5: Scheduled Jobs ---
    ws5 = wb.create_sheet("Scheduled Jobs")
    write_headers(ws5, ['Task Name', 'Path', 'State', 'Trigger', 'Command',
                        'Arguments', 'Last Run', 'Next Run'])
    for row, job in enumerate(selected_jobs, 2):
        ws5.cell(row=row, column=1, value=job.task_name)
        ws5.cell(row=row, column=2, value=job.task_path)
        ws5.cell(row=row, column=3, value=job.state)
        ws5.cell(row=row, column=4, value=job.trigger_info)
        ws5.cell(row=row, column=5, value=job.action_command)
        ws5.cell(row=row, column=6, value=job.action_arguments)
        ws5.cell(row=row, column=7, value=job.last_run)
        ws5.cell(row=row, column=8, value=job.next_run)

    # --- Tab 6: NSSM Services ---
    if nssm_services:
        ws6 = wb.create_sheet("NSSM Services")
        write_headers(ws6, ['Service Name', 'Display Name', 'Status', 'Startup',
                            'Application', 'Arguments', 'Working Dir', 'Related Project'])
        for row, svc in enumerate(nssm_services, 2):
            ws6.cell(row=row, column=1, value=svc.service_name)
            ws6.cell(row=row, column=2, value=svc.display_name)
            ws6.cell(row=row, column=3, value=svc.status)
            ws6.cell(row=row, column=4, value=svc.startup_type)
            ws6.cell(row=row, column=5, value=svc.application)
            ws6.cell(row=row, column=6, value=svc.app_arguments)
            ws6.cell(row=row, column=7, value=svc.app_directory)
            ws6.cell(row=row, column=8, value=svc.related_project)

    # --- Tab 7: SSMS Scripts ---
    if ssms_scripts:
        ws6 = wb.create_sheet("SSMS Scripts")
        write_headers(ws6, ['Filename', 'Lines', 'Modified', 'Tables Referenced',
                            'Schema Objects', 'Related Projects'])
        for row, s in enumerate(ssms_scripts, 2):
            ws6.cell(row=row, column=1, value=s.filename)
            ws6.cell(row=row, column=2, value=s.line_count)
            ws6.cell(row=row, column=3, value=s.modified[:10])
            ws6.cell(row=row, column=4, value=', '.join(s.tables_referenced[:10]))
            ws6.cell(row=row, column=5, value=', '.join(s.schema_objects[:5]))
            ws6.cell(row=row, column=6, value=', '.join(ssms_matches.get(s.filename, [])))

    # Auto-size columns (approximate)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    logger.info(f"Excel workbook saved: {output_path}")


# =============================================================================
# MASTER ORCHESTRATOR
# =============================================================================

def run_full_analysis(project_folders: List[str],
                      pycharm_root: str,
                      ssms_root: Optional[str],
                      selected_jobs: List[ScheduledJob],
                      output_root: str,
                      progress_callback=None) -> str:
    """
    Run the complete analysis pipeline.

    Args:
        project_folders: List of project folder names to scan
        pycharm_root: Path to PycharmProjects directory
        ssms_root: Path to SSMS folder (None to skip)
        selected_jobs: Task Scheduler jobs to include
        output_root: Base output directory (Results folder)
        progress_callback: Optional callable(message, percent) for GUI updates

    Returns:
        Path to the output directory
    """
    timestamp = datetime.now().strftime('%Y-%m-%d')
    output_dir = os.path.join(output_root, f"ProjectAnalyzer_{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

    prompts_dir = os.path.join(output_dir, "Prompts")
    os.makedirs(prompts_dir, exist_ok=True)

    def progress(msg, pct=0):
        logger.info(msg)
        if progress_callback:
            progress_callback(msg, pct)

    total_steps = len(project_folders) + 4  # +ssms +nssm +xref +export
    step = 0

    # --- Phase 1: Scan each project ---
    inventories = {}
    for folder_name in project_folders:
        step += 1
        pct = int(step / total_steps * 100)
        progress(f"Scanning {folder_name}...", pct)

        folder_path = os.path.join(pycharm_root, folder_name)
        if os.path.isdir(folder_path):
            inventories[folder_name] = scan_project(folder_path)
        else:
            logger.warning(f"Project folder not found: {folder_path}")

    # --- Phase 2: Scan SSMS ---
    step += 1
    ssms_scripts = []
    ssms_matches = {}
    if ssms_root:
        progress("Scanning SSMS scripts...", int(step / total_steps * 100))
        ssms_scripts = scan_ssms_folder(ssms_root)
        ssms_matches = match_ssms_to_projects(ssms_scripts, inventories)

    # --- Phase 2b: Scan NSSM services ---
    step += 1
    progress("Scanning NSSM services...", int(step / total_steps * 100))
    nssm_services = scan_nssm_services(pycharm_root)

    # --- Phase 3: Cross-reference ---
    step += 1
    progress("Building cross-project analysis...", int(step / total_steps * 100))
    xref = build_cross_reference(inventories)
    with open(os.path.join(output_dir, "xref.json"), "w", encoding="utf-8") as f:
         json.dump({
         "table_readers": xref.table_readers,
         "table_writers": xref.table_writers,
         "import_graph":  xref.import_graph,
    }, f, indent=2, default=list)

    # --- Phase 4: Generate outputs ---
    step += 1
    progress("Generating prompts and workbook...", int(step / total_steps * 100))

    # Per-project prompts
    for proj_name, inv in inventories.items():
        prompt = generate_project_prompt(inv, xref, selected_jobs,
                                         nssm_services, ssms_scripts, ssms_matches)
        prompt_file = os.path.join(prompts_dir, f"{proj_name}_Prompt.md")
        with open(prompt_file, 'w', encoding='utf-8') as f:
            f.write(prompt)

    # Excel workbook
    excel_path = os.path.join(output_dir, "ProjectAnalyzer_CrossReference.xlsx")
    export_workbook(inventories, xref, ssms_scripts, ssms_matches,
                    selected_jobs, nssm_services, excel_path)

    # Summary markdown
    summary_path = os.path.join(output_dir, "ProjectAnalyzer_Summary.md")
    _write_summary(inventories, xref, ssms_scripts, selected_jobs,
                   nssm_services, summary_path, timestamp)

    progress("Analysis complete!", 100)
    logger.info(f"Output written to: {output_dir}")

    return output_dir


def _write_summary(inventories, xref, ssms_scripts, selected_jobs,
                   nssm_services, output_path, timestamp):
    """Write the master summary markdown."""
    lines = []
    lines.append(f"# CRP Project Ecosystem Analysis")
    lines.append(f"_Generated: {timestamp}_")
    lines.append("")
    lines.append(f"## Overview")
    lines.append(f"- **Projects scanned:** {len(inventories)}")
    lines.append(f"- **Total scripts:** {sum(len(i.scripts) for i in inventories.values())}")
    lines.append(f"- **Total SQL files:** {sum(len(i.sql_files) for i in inventories.values())}")
    lines.append(f"- **Total documentation:** {sum(len(i.docs) for i in inventories.values())}")
    lines.append(f"- **Total lines of code:** {sum(i.total_lines for i in inventories.values()):,}")
    lines.append(f"- **SSMS scripts:** {len(ssms_scripts)}")
    lines.append(f"- **Scheduled jobs (Task Scheduler):** {len(selected_jobs)}")
    lines.append(f"- **Always-on services (NSSM):** {len(nssm_services)}")
    lines.append(f"- **Shared tables (multi-project):** {len(xref.shared_tables)}")
    lines.append("")

    # NSSM services summary
    if nssm_services:
        lines.append("## Always-On Services (NSSM)")
        for svc in nssm_services:
            proj = f" → {svc.related_project}" if svc.related_project else ""
            lines.append(f"- **{svc.service_name}** — {svc.status}, {svc.startup_type}{proj}")
            lines.append(f"  Application: {svc.application}")
        lines.append("")

    lines.append("## Projects")
    for name, inv in sorted(inventories.items()):
        lines.append(f"### {name}")
        lines.append(f"- {len(inv.scripts)} scripts, {inv.total_lines:,} lines")
        if inv.entry_points:
            lines.append(f"- Entry points: {', '.join(inv.entry_points)}")
        if inv.all_cross_imports:
            lines.append(f"- Depends on: {', '.join(sorted(inv.all_cross_imports))}")
        decision_logs = [d for d in inv.docs if d.doc_type == 'decision_log']
        if decision_logs:
            lines.append(f"- Decision logs: {', '.join(d.filename for d in decision_logs)}")
        # Flag if this project has an NSSM service
        proj_nssm = [s for s in nssm_services if s.related_project == name]
        if proj_nssm:
            lines.append(f"- NSSM service: {', '.join(s.service_name for s in proj_nssm)}")
        lines.append("")

    lines.append("## Shared Tables")
    for table, info in sorted(xref.shared_tables.items()):
        writers = ', '.join(info['writers'])
        readers = ', '.join(info['readers'])
        lines.append(f"- **{table}** — written by [{writers}], read by [{readers}]")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


# =============================================================================
# CLI ENTRY POINT
# =============================================================================

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='CRP Project Intelligence Analyzer')
    parser.add_argument('--projects', nargs='*', help='Specific projects to scan (default: all)')
    parser.add_argument('--no-ssms', action='store_true', help='Skip SSMS folder scan')
    parser.add_argument('--no-jobs', action='store_true', help='Skip Task Scheduler scan')
    parser.add_argument('--list', action='store_true', help='List discovered projects and exit')

    args = parser.parse_args()
    config = load_config()

    if args.list:
        projects = discover_projects(config['pycharm_root'])
        print(f"Found {len(projects)} projects in {config['pycharm_root']}:")
        for p in projects:
            print(f"  {p}")
        sys.exit(0)

    # Determine projects to scan
    if args.projects:
        project_list = args.projects
    else:
        project_list = discover_projects(config['pycharm_root'])

    # SSMS
    ssms_path = None if args.no_ssms else config.get('ssms_root', DEFAULT_SSMS_ROOT)

    # Jobs (CLI mode skips job selection — include none)
    jobs = [] if args.no_jobs else []  # GUI handles job selection

    output = run_full_analysis(
        project_folders=project_list,
        pycharm_root=config['pycharm_root'],
        ssms_root=ssms_path,
        selected_jobs=jobs,
        output_root=config.get('output_root', DEFAULT_OUTPUT_ROOT),
    )

    print(f"\nOutput: {output}")