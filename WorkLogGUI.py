"""
WorkLog.py - Monthly Work Activity Reconstructor
===================================================

Scans file modification dates and punchlist activity to reconstruct
a timeline of what was worked on during a given month. Designed to
support invoicing by combining:

  1. File modification timestamps from PycharmProjects and SSMS folders
  2. PMA_PunchlistItems date activity (Created, Modified, Completed)
  3. Session clustering with estimated durations

Exports to xlsx with three sheets: Sessions, File Activity, Punchlist Activity.

Lives in: CRPUtils folder
Output:   WorkLog_YYYY-MM.xlsx in CRPUtils folder

Author: Pat Yearick
Created: March 2026
"""

import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from pathlib import Path
import logging

# =============================================================================
# CONFIGURATION
# =============================================================================

SQL_SERVER = "BI-SQL001"
SQL_DATABASE = "CRPAF"
SQL_DRIVER = "ODBC Driver 17 for SQL Server"
LOG_FILE = r"C:\Logs\worklog.log"

SCAN_DIRS = [
    r"C:\Users\pyearick.CRP\OneDrive - CRP Industries Inc\CRPAF\PycharmProjects",
    r"C:\Users\pyearick.CRP\OneDrive - CRP Industries Inc\CRPAF\SQL Server Management Studio",
    r"C:\Users\pyearick.CRP\Documents\Snagit",
]

SKIP_FOLDERS = {
    '.idea', '.git', '.venv', '__pycache__', 'node_modules',
    'CommitsGH', '.ipynb_checkpoints', '.venvBISQL001', 'Archive'
}

# Files modified within this many minutes of each other = same session
SESSION_GAP_MINUTES = 30

OUTPUT_DIR = r"C:\Users\pyearick.CRP\OneDrive - CRP Industries Inc\CRPAF\PycharmProjects\CRPUtils"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# =============================================================================
# DATA COLLECTION
# =============================================================================

def get_connection():
    import pyodbc
    return pyodbc.connect(
        f"DRIVER={{{SQL_DRIVER}}};"
        f"SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DATABASE};"
        "Trusted_Connection=yes;"
    )


def scan_files(year, month):
    """
    Scan configured directories for files modified in the given month.
    Top-level only: files directly in each scan dir + files directly
    inside immediate child folders (project folders). No deeper recursion.
    """
    from calendar import monthrange
    start = datetime(year, month, 1)
    _, last_day = monthrange(year, month)
    end = datetime(year, month, last_day, 23, 59, 59)

    results = []

    for scan_dir in SCAN_DIRS:
        base = Path(scan_dir)
        if not base.exists():
            logger.warning(f"Directory not found: {scan_dir}")
            continue

        source_label = base.name  # "PycharmProjects" or "SQL Server Management Studio"

        # Files directly in the scan dir
        for f in base.iterdir():
            if f.is_file():
                _check_file(f, start, end, source_label, "(root)", results)

        # Files directly inside immediate child folders
        for child in sorted(base.iterdir()):
            if not child.is_dir():
                continue
            if child.name in SKIP_FOLDERS or child.name.startswith('.'):
                continue

            project_name = child.name
            for f in child.iterdir():
                if f.is_file():
                    _check_file(f, start, end, source_label, project_name, results)

    results.sort(key=lambda r: r['modified'])
    logger.info(f"File scan: {len(results)} files modified in {year}-{month:02d}")
    return results


def _check_file(filepath, start, end, source, project, results):
    """Check if a file was modified in the date range and add to results."""
    try:
        mtime = datetime.fromtimestamp(filepath.stat().st_mtime)
        if start <= mtime <= end:
            results.append({
                'source': source,
                'project': project,
                'filename': filepath.name,
                'modified': mtime,
                'size_kb': round(filepath.stat().st_size / 1024, 1),
                'extension': filepath.suffix.lower(),
            })
    except (OSError, ValueError):
        pass


def fetch_punchlist_activity(year, month):
    """
    Fetch punchlist items with any date activity in the given month.
    An item qualifies if CreatedDate, LastModifiedDate, or CompletedDate
    falls within the month.
    """
    from calendar import monthrange
    start = datetime(year, month, 1)
    _, last_day = monthrange(year, month)
    end = datetime(year, month, last_day, 23, 59, 59)

    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        SELECT PunchlistItemID, Project, ItemNumber, Title, Status, Priority,
               CreatedDate, LastModifiedDate, CompletedDate
        FROM [dbo].[PMA_PunchlistItems]
        WHERE (CreatedDate BETWEEN ? AND ?)
           OR (LastModifiedDate BETWEEN ? AND ?)
           OR (CompletedDate BETWEEN ? AND ?)
        ORDER BY
            COALESCE(LastModifiedDate, CreatedDate, CompletedDate)
    """

    cursor.execute(sql, (start, end, start, end, start, end))
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    cursor.close()
    conn.close()

    items = [dict(zip(columns, row)) for row in rows]
    logger.info(f"Punchlist query: {len(items)} items with activity in {year}-{month:02d}")
    return items


# =============================================================================
# SESSION CLUSTERING
# =============================================================================

def cluster_sessions(file_results):
    """
    Group file modifications into work sessions.
    Files in the same project modified within SESSION_GAP_MINUTES of each
    other are considered one session. Returns a list of session dicts.
    """
    if not file_results:
        return []

    # Sort by project then timestamp
    sorted_files = sorted(file_results, key=lambda r: (r['project'], r['modified']))

    sessions = []
    current = None

    for f in sorted_files:
        if (current is None
                or f['project'] != current['project']
                or (f['modified'] - current['last_touch']) > timedelta(minutes=SESSION_GAP_MINUTES)):
            # Start a new session
            if current:
                sessions.append(_finalize_session(current))
            current = {
                'project': f['project'],
                'source': f['source'],
                'start': f['modified'],
                'last_touch': f['modified'],
                'files': [f['filename']],
                'extensions': {f['extension']},
            }
        else:
            current['last_touch'] = f['modified']
            current['files'].append(f['filename'])
            current['extensions'].add(f['extension'])

    if current:
        sessions.append(_finalize_session(current))

    sessions.sort(key=lambda s: (s['date'], s['start_time']))
    logger.info(f"Clustering: {len(sessions)} sessions from {len(file_results)} files")
    return sessions


def _finalize_session(session_data):
    """Calculate duration and build the session summary."""
    duration = session_data['last_touch'] - session_data['start']
    duration_mins = max(int(duration.total_seconds() / 60), 1)

    # If only 1 file, estimate ~5 min minimum
    if len(session_data['files']) == 1:
        duration_mins = max(duration_mins, 5)

    unique_files = sorted(set(session_data['files']))

    return {
        'date': session_data['start'].strftime('%Y-%m-%d'),
        'start_time': session_data['start'].strftime('%H:%M'),
        'end_time': session_data['last_touch'].strftime('%H:%M'),
        'duration_mins': duration_mins,
        'source': session_data['source'],
        'project': session_data['project'],
        'file_count': len(session_data['files']),
        'unique_files': len(unique_files),
        'file_types': ', '.join(sorted(session_data['extensions'])),
        'sample_files': '; '.join(unique_files[:5]),
    }


# =============================================================================
# XLSX EXPORT
# =============================================================================

def export_xlsx(year, month, sessions, file_results, punchlist_items):
    """Write the three-sheet workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(name='Arial', bold=True, size=10)
    header_fill = PatternFill('solid', fgColor='2C3E50')
    header_font_white = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    data_font = Font(name='Arial', size=9)
    date_fmt = 'MM/DD/YYYY HH:MM'
    short_date_fmt = 'MM/DD/YYYY'
    thin_border = Border(
        bottom=Side(style='thin', color='D5D8DC')
    )

    def style_header(ws, headers, col_widths):
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[cell.column_letter].width = width
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    # --- Sheet 1: Sessions ---
    ws_sessions = wb.active
    ws_sessions.title = "Sessions"

    session_headers = [
        'Date', 'Start', 'End', 'Duration (min)', 'Source',
        'Project', 'Files', 'Unique Files', 'File Types', 'Sample Files'
    ]
    session_widths = [12, 8, 8, 14, 24, 20, 8, 12, 14, 50]
    style_header(ws_sessions, session_headers, session_widths)

    for row_idx, s in enumerate(sessions, 2):
        ws_sessions.cell(row=row_idx, column=1, value=s['date']).font = data_font
        ws_sessions.cell(row=row_idx, column=2, value=s['start_time']).font = data_font
        ws_sessions.cell(row=row_idx, column=3, value=s['end_time']).font = data_font
        ws_sessions.cell(row=row_idx, column=4, value=s['duration_mins']).font = data_font
        ws_sessions.cell(row=row_idx, column=5, value=s['source']).font = data_font
        ws_sessions.cell(row=row_idx, column=6, value=s['project']).font = data_font
        ws_sessions.cell(row=row_idx, column=7, value=s['file_count']).font = data_font
        ws_sessions.cell(row=row_idx, column=8, value=s['unique_files']).font = data_font
        ws_sessions.cell(row=row_idx, column=9, value=s['file_types']).font = data_font
        ws_sessions.cell(row=row_idx, column=10, value=s['sample_files']).font = data_font

        for col in range(1, 11):
            ws_sessions.cell(row=row_idx, column=col).border = thin_border

    # --- Sheet 2: File Activity ---
    ws_files = wb.create_sheet("File Activity")

    file_headers = ['Modified', 'Source', 'Project', 'Filename', 'Extension', 'Size (KB)']
    file_widths = [18, 24, 20, 40, 10, 10]
    style_header(ws_files, file_headers, file_widths)

    for row_idx, f in enumerate(file_results, 2):
        cell = ws_files.cell(row=row_idx, column=1, value=f['modified'])
        cell.font = data_font
        cell.number_format = date_fmt
        ws_files.cell(row=row_idx, column=2, value=f['source']).font = data_font
        ws_files.cell(row=row_idx, column=3, value=f['project']).font = data_font
        ws_files.cell(row=row_idx, column=4, value=f['filename']).font = data_font
        ws_files.cell(row=row_idx, column=5, value=f['extension']).font = data_font
        ws_files.cell(row=row_idx, column=6, value=f['size_kb']).font = data_font

        for col in range(1, 7):
            ws_files.cell(row=row_idx, column=col).border = thin_border

    # --- Sheet 3: Punchlist Activity ---
    ws_punch = wb.create_sheet("Punchlist Activity")

    punch_headers = [
        'Project', 'Item #', 'Title', 'Status', 'Priority',
        'Created', 'Modified', 'Completed'
    ]
    punch_widths = [16, 12, 50, 14, 10, 18, 18, 18]
    style_header(ws_punch, punch_headers, punch_widths)

    for row_idx, item in enumerate(punchlist_items, 2):
        ws_punch.cell(row=row_idx, column=1, value=item['Project']).font = data_font
        ws_punch.cell(row=row_idx, column=2, value=item['ItemNumber']).font = data_font
        ws_punch.cell(row=row_idx, column=3, value=item['Title'][:80]).font = data_font
        ws_punch.cell(row=row_idx, column=4, value=item['Status']).font = data_font
        ws_punch.cell(row=row_idx, column=5, value=item['Priority']).font = data_font

        for col_idx, date_key in [(6, 'CreatedDate'), (7, 'LastModifiedDate'), (8, 'CompletedDate')]:
            dt_val = item.get(date_key)
            cell = ws_punch.cell(row=row_idx, column=col_idx, value=dt_val)
            cell.font = data_font
            if dt_val:
                cell.number_format = date_fmt

        for col in range(1, 9):
            ws_punch.cell(row=row_idx, column=col).border = thin_border

    # --- Save ---
    filename = f"WorkLog_{year}-{month:02d}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(output_path)
    logger.info(f"Exported: {output_path}")
    return output_path


# =============================================================================
# GUI
# =============================================================================

COLORS = {
    'bg': '#f5f5f5',
    'header_bg': '#2c3e50',
    'header_fg': '#ffffff',
    'btn_primary': '#3498db',
    'btn_success': '#27ae60',
    'btn_fg': '#ffffff',
}


class WorkLogApp:
    def __init__(self, root):
        self.root = root
        self.root.title("WorkLog — Monthly Activity Export")
        self.root.geometry("500x340")
        self.root.configure(bg=COLORS['bg'])
        self.root.resizable(False, False)

        self._build_header()
        self._build_controls()
        self._build_status()

    def _build_header(self):
        header = tk.Frame(self.root, bg=COLORS['header_bg'], height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(
            header, text="WorkLog — Monthly Activity Export",
            font=('Segoe UI', 14, 'bold'),
            bg=COLORS['header_bg'], fg=COLORS['header_fg']
        ).pack(side=tk.LEFT, padx=15)

    def _build_controls(self):
        frame = tk.LabelFrame(
            self.root, text="Select Month",
            font=('Segoe UI', 10, 'bold'),
            bg=COLORS['bg'], padx=15, pady=10
        )
        frame.pack(fill=tk.X, padx=20, pady=15)

        row1 = tk.Frame(frame, bg=COLORS['bg'])
        row1.pack(fill=tk.X, pady=5)

        now = datetime.now()

        tk.Label(row1, text="Year:", bg=COLORS['bg'],
                 font=('Segoe UI', 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.year_var = tk.StringVar(value=str(now.year))
        year_spin = ttk.Spinbox(row1, from_=2024, to=2030,
                                textvariable=self.year_var, width=6,
                                font=('Segoe UI', 10))
        year_spin.pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(row1, text="Month:", bg=COLORS['bg'],
                 font=('Segoe UI', 10)).pack(side=tk.LEFT, padx=(0, 5))
        months = [
            '01 - January', '02 - February', '03 - March',
            '04 - April', '05 - May', '06 - June',
            '07 - July', '08 - August', '09 - September',
            '10 - October', '11 - November', '12 - December',
        ]
        self.month_var = tk.StringVar(value=months[now.month - 1])
        month_combo = ttk.Combobox(row1, textvariable=self.month_var,
                                   values=months, state='readonly', width=16,
                                   font=('Segoe UI', 10))
        month_combo.pack(side=tk.LEFT)

        # Buttons
        btn_frame = tk.Frame(frame, bg=COLORS['bg'])
        btn_frame.pack(fill=tk.X, pady=(15, 5))

        self._make_button(btn_frame, "Generate WorkLog", self.run_export,
                          COLORS['btn_success']).pack(side=tk.LEFT, padx=5)

        self.open_var = tk.BooleanVar(value=True)
        tk.Checkbutton(btn_frame, text="Open file after export",
                       variable=self.open_var, bg=COLORS['bg'],
                       font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=15)

    def _build_status(self):
        self.status_frame = tk.LabelFrame(
            self.root, text="Status",
            font=('Segoe UI', 10, 'bold'),
            bg=COLORS['bg'], padx=10, pady=5
        )
        self.status_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))

        self.status_label = tk.Label(
            self.status_frame, text="Select a month and click Generate.",
            bg=COLORS['bg'], font=('Segoe UI', 9),
            anchor='w', justify=tk.LEFT, wraplength=440
        )
        self.status_label.pack(fill=tk.X, pady=5)

    def _make_button(self, parent, text, command, color):
        return tk.Button(
            parent, text=text, command=command,
            bg=color, fg=COLORS['btn_fg'],
            font=('Segoe UI', 10, 'bold'),
            relief=tk.FLAT, padx=15, pady=5,
            activebackground=color, activeforeground=COLORS['btn_fg'],
            cursor='hand2'
        )

    def _set_status(self, msg):
        self.status_label.config(text=msg)
        self.root.update_idletasks()

    def _get_month_year(self):
        year = int(self.year_var.get())
        month = int(self.month_var.get().split(' ')[0])
        return year, month

    def run_export(self):
        year, month = self._get_month_year()
        month_label = f"{year}-{month:02d}"

        try:
            # Step 1: Scan files
            self._set_status(f"Scanning files for {month_label}...")
            file_results = scan_files(year, month)

            # Step 2: Query punchlist
            self._set_status(f"Querying punchlist activity for {month_label}...")
            punchlist_items = fetch_punchlist_activity(year, month)

            # Step 3: Cluster sessions
            self._set_status("Clustering work sessions...")
            sessions = cluster_sessions(file_results)

            # Step 4: Export
            self._set_status("Writing xlsx...")
            output_path = export_xlsx(year, month, sessions, file_results, punchlist_items)

            summary = (
                f"Exported: {os.path.basename(output_path)}\n"
                f"Sessions: {len(sessions)}  |  "
                f"Files: {len(file_results)}  |  "
                f"Punchlist items: {len(punchlist_items)}"
            )
            self._set_status(summary)

            if self.open_var.get():
                os.startfile(output_path)

        except Exception as e:
            logger.error(f"Export failed: {e}", exc_info=True)
            messagebox.showerror("Export Error", f"Failed:\n{e}")
            self._set_status(f"Error: {e}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = WorkLogApp(root)
    root.mainloop()