"""
excel_format.py — shared Excel export formatting for CRPAF workbooks
=============================================================================
ONE home for the CRPAF workbook-export convention so projects stop copy-pasting
their own `autofit_ws`. Every tab in an exported .xlsx should be readable at a
glance: columns auto-sized to their content (capped so one long free-text field
can't blow the sheet out) and the header row frozen.

Standing convention (see CRPAF/PycharmProjects/CLAUDE.md): autosize every sheet
to a max of 100 characters and freeze the header row.

Usage — inside a pandas ExcelWriter (the common case):
    from CRPUtils.excel_format import autofit_worksheet
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="Data", index=False)
        autofit_worksheet(xl.sheets["Data"], df)      # widths from the frame

Or on any openpyxl worksheet with no DataFrame handy:
    autofit_worksheet(ws)                             # widths from cell values

NOTE: run this on DATA sheets BEFORE stamping any native Excel PivotTable via
CRPUtils.excel_pivot — an openpyxl re-read/write can disturb the pivot cache.
The pivot helper sets its own label-column width.
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter

DEFAULT_MAX_WIDTH = 100   # CRPAF standing cap, in characters
_PAD = 2                  # a little breathing room past the longest value


def autofit_worksheet(ws, df=None, max_width=DEFAULT_MAX_WIDTH, freeze_header=True):
    """Auto-size each column to max(header, longest cell) + 2, capped at
    max_width, and (by default) freeze the header row.

    ws : an openpyxl worksheet (e.g. ``xl.sheets[name]`` from a pandas
         ExcelWriter, or ``wb[name]``).
    df : the DataFrame written to this sheet. If given, widths come from it
         (fast, vectorized, and matches exactly what was written). If None,
         widths are derived from the worksheet's own cell values.
    max_width : hard cap in characters (CRPAF default 100) so one long
         free-text column can't blow the sheet out.
    freeze_header : freeze the top row so headers stay visible when scrolling.

    Returns the worksheet, so callers can chain if they like.
    """
    if df is not None:
        import pandas as pd
        for i, col in enumerate(df.columns, start=1):
            longest = df[col].astype(str).str.len().max()   # NaN on an empty frame
            longest = 0 if pd.isna(longest) else int(longest)
            width = max(len(str(col)), longest) + _PAD
            ws.column_dimensions[get_column_letter(i)].width = min(width, max_width)
    else:
        for i, cells in enumerate(ws.columns, start=1):
            longest = 0
            for c in cells:
                if c.value is not None:
                    n = len(str(c.value))
                    if n > longest:
                        longest = n
            ws.column_dimensions[get_column_letter(i)].width = min(longest + _PAD, max_width)

    if freeze_header:
        ws.freeze_panes = "A2"
    return ws
